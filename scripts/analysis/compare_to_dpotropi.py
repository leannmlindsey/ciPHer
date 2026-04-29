"""Compute Recall@N for cipher's K head on the DpoTropiSearch in-vitro
test set, and compare against their published TropiGAT / TropiSEQ /
SpikeHunter / Random Recall@N curves.

DpoTropiSearch (Concho-Eloko et al., Nat Commun 2025) reports Recall@N
on a curated test set of 240 phage-protein-target rows from in-vitro
validated lytic depolymerases (PhageHostLearn-derived). Their
ground-truth, predictions, and reported metric values all live in the
Zenodo supplements:

  Supplementary_Data_5  — per-protein ground truth + per-model predictions
  source_figure4A       — published Recall@N curves per model

Their metric definition (recovered from B.Compile_results.ipynb):

  For each (phage_protein, true_KL) row:
      hit = 1 if true_KL ∈ model's top-N predictions for that protein,
            else 0
  Recall@N = sum(hits) / total rows

We replicate this exactly for cipher's K head:
  1. Load Supplementary_Data_5 → DataFrame.
  2. Load DpoTropi's pre-computed ESM-2 650M embeddings (1280-dim,
     matching cipher's LAPTOP repro / ESM-2 sweep models).
  3. For each protein, run cipher's K head, take top-N K classes.
  4. Canonicalize K↔KL (cipher's "K1" ≡ DpoTropi's "KL1"; cipher's
     "KL101" already matches).
  5. Compute Recall@N for N=1..40.
  6. Stack against their reported numbers and emit a CSV + plot.

Caveats:
  - Cipher's LAPTOP repro K head is trained on ESM-2 650M MEAN. Their
    embeddings are also ESM-2 650M MEAN, so direct comparison is valid.
  - Cipher's K vocabulary covers 161 classes vs DpoTropi's ~117. A few
    KL types may be out-of-vocab for cipher → guaranteed misses there.
  - DpoTropi's "TropiGAT & TropiSeq" combined number is set-union of
    top-N predictions from both models, then recall over the union.
    We don't need to reproduce that; comparing cipher to each model
    individually is sufficient.

Output:
  results/dpotropi_comparison.csv  (Recall@N table)
  results/figures/dpotropi_recall_at_n.svg

Usage:
  PYTHONPATH=src python3 scripts/analysis/compare_to_dpotropi.py \\
      experiments/attention_mlp/repro_old_v3_in_cipher_LAPTOP_20260425_235817
"""

import argparse
import csv
import importlib.util
import json
import os
import re
import sys
from pathlib import Path

import numpy as np


# Cipher importability
def _ensure_cipher_on_path():
    here = Path(__file__).resolve().parent
    for parent in [here, *here.parents]:
        cand = parent / 'src' / 'cipher'
        if cand.is_dir():
            src = str(parent / 'src')
            if src not in sys.path:
                sys.path.insert(0, src)
            return

_ensure_cipher_on_path()


def load_supp5(xlsx_path):
    """Return list of dicts: phage, protein, target (set of KL types), folds."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb['Supplementary_Data_5']
    rows = list(ws.iter_rows(values_only=True))
    # Header is row 2 (0-indexed); data starts row 3
    header = [str(c).strip() if c else '' for c in rows[2]]
    out = []
    for r in rows[3:]:
        if not r or not r[0]:
            continue
        d = dict(zip(header, r))
        target_str = (d.get('Targets') or '').strip()
        if not target_str:
            continue
        # Targets can be comma- or semicolon-separated, sometimes both
        targets = set()
        for t in re.split(r'[;,]', target_str):
            t = t.strip()
            if t and t.lower() not in ('no_hits', 'none', ''):
                targets.add(t)
        if not targets:
            continue
        out.append({
            'phage': d.get('phage'),
            'protein': d.get('protein'),
            'fold': d.get('Folds'),
            'targets': targets,
        })
    return out


def load_dpotropi_embeddings(csv_path):
    """{protein_id: 1280-dim ndarray} from their CSV."""
    out = {}
    with open(csv_path) as f:
        rdr = csv.reader(f)
        for row in rdr:
            pid = row[0]
            vec = np.array([float(x) for x in row[1:]], dtype=np.float32)
            out[pid] = vec
    return out


def build_phage_to_prokka_prefix(phage_protein_mapping_csv):
    """Cipher PHL: phage A1a → Prokka prefix DGCOCGBI (from DGCOCGBI_00014).

    Returns {phage_id: prokka_prefix}. Each phage's proteins all share the
    same Prokka prefix (Prokka assigns one 8-letter hash per genome).
    """
    out = {}
    with open(phage_protein_mapping_csv) as f:
        rdr = csv.DictReader(f)
        for row in rdr:
            phage = row['matrix_phage_name']
            prot = row['protein_id']
            if '_' in prot:
                prefix = prot.rsplit('_', 1)[0]
                if phage not in out:
                    out[phage] = prefix
    return out


def dpotropi_to_cipher_pid(phage, dpotropi_protein_id, phage_to_prefix):
    """Translate A1a_00014 → DGCOCGBI_00014 (cipher's protein_id form)."""
    if '_' not in dpotropi_protein_id:
        return None
    suffix = dpotropi_protein_id.rsplit('_', 1)[1]  # "00014"
    prefix = phage_to_prefix.get(phage)
    if prefix is None:
        return None
    return f'{prefix}_{suffix}'


def load_cipher_predictor(experiment_dir):
    from cipher.evaluation.runner import find_predict_module
    predict_path, model_dir = find_predict_module(experiment_dir)
    if predict_path is None:
        raise FileNotFoundError(f'No predict.py for {experiment_dir}')
    if model_dir not in sys.path:
        sys.path.insert(0, model_dir)
    spec = importlib.util.spec_from_file_location('predict', predict_path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod.get_predictor(os.path.abspath(experiment_dir))


def to_kl(name):
    """Canonicalize cipher class name to DpoTropi KL form.

    K1 -> KL1
    KL101 -> KL101 (unchanged)
    null / N/A -> None
    """
    if name in (None, 'null', 'N/A', ''):
        return None
    if name.startswith('KL'):
        return name
    if name.startswith('K') and name[1:].isdigit():
        return 'KL' + name[1:]
    return name  # other (rare)


def topn_predictions(predictor, embedding, n):
    """Return list of top-N K class names (canonicalized to KL form),
    descending by probability."""
    out = predictor.predict_protein(embedding)
    k_probs = out.get('k_probs', {})
    sorted_classes = sorted(k_probs.items(), key=lambda kv: -kv[1])
    return [to_kl(c) for c, _ in sorted_classes[:n] if to_kl(c) is not None]


def load_published_recall_at_n(xlsx_path):
    """source_figure4A → dict {N: {model: recall_pct}}."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb['source_figure4A']
    rows = list(ws.iter_rows(values_only=True))
    # Find header row (has 'Top N')
    hdr_i = next((i for i, r in enumerate(rows)
                  if r and any(c == 'Top N' for c in r if c)), None)
    if hdr_i is None:
        return {}
    header = list(rows[hdr_i])
    # Identify column indices
    n_idx = header.index('Top N')
    cols = {}
    for i, h in enumerate(header):
        if h and h != 'Top N' and i != 0:
            cols[h] = i
    out = {}
    for r in rows[hdr_i + 1:]:
        if not r or r[n_idx] is None:
            continue
        try:
            n = int(float(r[n_idx]))
        except (TypeError, ValueError):
            continue
        out[n] = {model: float(r[i]) for model, i in cols.items()
                  if r[i] is not None}
    return out


def main():
    p = argparse.ArgumentParser()
    p.add_argument('experiment_dir',
                   help='cipher experiment dir (must have model_k/, ESM-2 650M)')
    p.add_argument('--dpotropi-repo',
                   default='/Users/leannmlindsey/WORK/CLAUDE_ELOKO/DpoTropiSearch')
    p.add_argument('--zenodo-dir',
                   default='/Users/leannmlindsey/WORK/PHI_TSP/DpoTropiSearch_zenoto_data')
    p.add_argument('--max-n', type=int, default=40)
    p.add_argument('--out-csv',
                   default='results/dpotropi_comparison.csv')
    p.add_argument('--out-svg',
                   default='results/figures/dpotropi_recall_at_n.svg')
    args = p.parse_args()

    print('Loading DpoTropi data ...')
    supp5 = load_supp5(os.path.join(args.zenodo_dir,
                                     '41467_2025_63861_MOESM3_ESM.xlsx'))
    print(f'  Supplementary_Data_5 rows with targets: {len(supp5)}')
    published = load_published_recall_at_n(
        os.path.join(args.zenodo_dir,
                     '41467_2025_63861_MOESM6_ESM.xlsx'))
    print(f'  published Recall@N rows: {len(published)} '
          f'(N range: {min(published)}..{max(published)})')

    # DpoTropi protein IDs (e.g. A1a_00014) translate to cipher PHL
    # protein IDs (e.g. DGCOCGBI_00014) via the phage→Prokka-prefix map
    # in cipher's PHL phage_protein_mapping.csv.
    phl_pp_map = ('data/validation_data/HOST_RANGE/PhageHostLearn/'
                  'metadata/phage_protein_mapping.csv')
    phage_to_prefix = build_phage_to_prokka_prefix(phl_pp_map)
    print(f'  phage→prokka_prefix map: {len(phage_to_prefix)} phages')

    # Use cipher's existing PHL ESM-2 embeddings (keyed by MD5 of protein
    # sequence). We need pid→md5 from the PHL FASTA, then md5→embedding
    # from the val NPZ.
    from cipher.data.embeddings import load_embeddings
    from cipher.data.proteins import load_fasta_md5
    print('  Loading cipher PHL pid→md5 + ESM-2 mean embeddings ...')
    pid_md5 = load_fasta_md5('data/validation_data/metadata/validation_rbps_all.faa')
    val_emb_path = ('/Users/leannmlindsey/WORK/PHI_TSP/phi_tsp/klebsiella/'
                    'validation_data/combined/validation_inputs/'
                    'validation_embeddings_md5.npz')
    emb = load_embeddings(val_emb_path)
    print(f'    pid_md5: {len(pid_md5)} proteins, '
          f'embeddings: {len(emb)} MD5s')

    print(f'\nLoading cipher predictor: {args.experiment_dir}')
    pred = load_cipher_predictor(args.experiment_dir)
    cipher_K_classes = set(to_kl(c) for c in (pred.k_classes or [])
                           if to_kl(c) is not None)
    print(f'  cipher K vocabulary (canonicalized): {len(cipher_K_classes)} KL types')

    # KL-vocab coverage check
    all_targets = set()
    for r in supp5:
        all_targets.update(r['targets'])
    in_vocab = all_targets & cipher_K_classes
    oov = all_targets - cipher_K_classes
    print(f'  DpoTropi test KL types: {len(all_targets)}')
    print(f'    in cipher vocab:  {len(in_vocab)}')
    print(f'    OOV for cipher:   {len(oov)}')
    if oov:
        print(f'    OOV examples:    {sorted(oov)[:10]}')

    # Translate each DpoTropi row to cipher pid + embedding lookup
    n_no_phage = 0
    n_no_pid = 0
    n_no_md5 = 0
    n_no_emb = 0
    matched = []
    for r in supp5:
        if r['phage'] not in phage_to_prefix:
            n_no_phage += 1
            continue
        cipher_pid = dpotropi_to_cipher_pid(
            r['phage'], r['protein'], phage_to_prefix)
        if cipher_pid is None:
            n_no_pid += 1
            continue
        md5 = pid_md5.get(cipher_pid)
        if md5 is None:
            n_no_md5 += 1
            continue
        if md5 not in emb:
            n_no_emb += 1
            continue
        matched.append({**r, 'cipher_pid': cipher_pid,
                        'embedding': emb[md5]})
    print(f'\n  rows matched to cipher embeddings: {len(matched)} / {len(supp5)}')
    print(f'    skipped (no phage-prefix):   {n_no_phage}')
    print(f'    skipped (no cipher pid):     {n_no_pid}')
    print(f'    skipped (no md5 in fasta):   {n_no_md5}')
    print(f'    skipped (no embedding):      {n_no_emb}')

    # Compute cipher top-N predictions per row
    print('\nRunning cipher K head ...')
    rows_with_pred = []
    for r in matched:
        topN = topn_predictions(pred, r['embedding'], args.max_n)
        rows_with_pred.append({**r, 'cipher_topN': topN})

    # Recall@N for cipher: per N, fraction of rows where any of the row's
    # targets is in cipher's top-N.
    cipher_recall = {}
    for N in range(1, args.max_n + 1):
        hits = sum(
            1 for r in rows_with_pred
            if any(t in r['cipher_topN'][:N] for t in r['targets'])
        )
        cipher_recall[N] = 100.0 * hits / max(len(rows_with_pred), 1)

    # Build comparison table
    out_dir = os.path.dirname(args.out_csv) or '.'
    os.makedirs(out_dir, exist_ok=True)
    model_cols = sorted(set(m for d in published.values() for m in d))
    model_cols = [m for m in model_cols if m != 'Top N']
    with open(args.out_csv, 'w', newline='') as f:
        w = csv.writer(f)
        header = ['N', 'cipher_LAPTOP_repro_ESM2'] + model_cols
        w.writerow(header)
        for N in range(1, args.max_n + 1):
            row = [N, f'{cipher_recall[N]:.2f}']
            d = published.get(N, {})
            for m in model_cols:
                row.append(f'{d.get(m, ""):.1f}' if d.get(m) is not None else '')
            w.writerow(row)
    print(f'\nWrote {args.out_csv}')

    # Print headline rows
    print()
    print(f'{"N":>3}  {"cipher":>7}  ' + '  '.join(f'{m[:14]:>14}'
                                                    for m in model_cols))
    print('-' * (12 + 16 * len(model_cols)))
    for N in (1, 3, 5, 10, 20, 40):
        if N > args.max_n: continue
        d = published.get(N, {})
        cells = [f'{d.get(m, 0):>14.1f}' if d.get(m) is not None else f'{"-":>14}'
                 for m in model_cols]
        print(f'{N:>3}  {cipher_recall[N]:>7.2f}  ' + '  '.join(cells))

    # Plot
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        Ns = list(range(1, args.max_n + 1))
        fig, ax = plt.subplots(figsize=(8, 5))
        ax.plot(Ns, [cipher_recall[N] for N in Ns],
                color='#1f77b4', lw=2.5, marker='o', markersize=4,
                label='cipher (LAPTOP, ESM-2 mean)')
        colors = {'TropiGAT': '#d62728', 'TropiSeq': '#2ca02c',
                  'TropiGAT & TropiSeq': '#9467bd', 'SpikeHunter': '#ff7f0e',
                  'Random appraoch': '#999999', 'Random approach': '#999999'}
        for m in model_cols:
            ys = [published.get(N, {}).get(m) for N in Ns]
            ys = [y for y in ys if y is not None]
            if not ys: continue
            ax.plot(Ns[:len(ys)], ys,
                    color=colors.get(m, '#888'), lw=1.8, marker='s',
                    markersize=3, label=m,
                    linestyle='--' if 'Random' in m else '-')
        ax.set_xlabel('Top N')
        ax.set_ylabel('Recall@N (%)')
        ax.set_title(f'Recall@N on DpoTropi in-vitro test set\n'
                     f'(n={len(rows_with_pred)} phage-protein rows)',
                     fontsize=11, fontweight='bold')
        ax.set_xlim(0.5, args.max_n + 0.5)
        ax.set_ylim(0, max(105, max(cipher_recall.values()) + 5))
        ax.set_xticks([1, 5, 10, 15, 20, 25, 30, 35, 40][:1 + args.max_n // 5])
        ax.grid(alpha=0.3)
        ax.legend(loc='lower right', fontsize=8)
        os.makedirs(os.path.dirname(args.out_svg) or '.', exist_ok=True)
        fig.tight_layout()
        fig.savefig(args.out_svg, format='svg', bbox_inches='tight')
        fig.savefig(args.out_svg.replace('.svg', '.png'),
                    dpi=150, bbox_inches='tight')
        plt.close(fig)
        print(f'Wrote {args.out_svg}')
    except ImportError:
        print('matplotlib not available; skipping plot')


if __name__ == '__main__':
    main()
