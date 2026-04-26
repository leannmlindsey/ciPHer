"""Run OLD class-ranking and NEW host-ranking on the SAME trained model
and emit one row per positive (phage, host) pair with both methodologies'
predictions side-by-side.

Why this exists
---------------
We confirmed via `diff_trained_weights.py` that the OLD klebsiella model
and the ciPHer port produce BIT-EXACT identical weights when trained
with the same recipe on the same hardware. Yet PHL HR@1 differs:
  - OLD eval: ~0.291  (ranks K + O classes merged: "is host's K or O at rank 1 in some RBP?")
  - NEW eval: ~0.153  (ranks ~200 candidate hosts: "is the actual host at rank 1?")

The two metrics measure different things. This script makes that
concrete by running both on the same predictions.

Output columns (per (phage, host) positive pair):

  Identifiers:
    dataset, phage_id, host_id, true_K, true_O, n_rbps_with_emb

  OLD methodology — raw class ranking (matches old --merge-strategy raw):
    old_merged_rank          rank of first K==true_K or O==true_O
    old_top1_class           top-1 class (across K+O merged) for best RBP
    old_top1_prob            top-1 probability
    old_top1_head            'K' or 'O'
    old_top1_correct         True iff old_top1_class == true_K or true_O
    old_best_rbp             the RBP id whose top-1 produced old_merged_rank
    old_k_rank, old_o_rank   per-head ranks for diagnostics

  NEW methodology — host ranking (matches cipher.evaluation.runner):
    new_host_rank            competition rank of this host among candidates
    new_top1_host            top-1 ranked host id
    new_top1_host_K          its K type
    new_top1_host_O          its O type
    new_top1_score           its score (max over RBPs of max(K_p[host_K], O_p[host_O]))
    new_top1_correct         True iff new_top1_host == host_id (this row)

Usage
-----
    python scripts/analysis/compare_eval_methodologies.py \\
        experiments/attention_mlp/repro_old_v3_in_cipher_LAPTOP_<TS>/ \\
        --datasets PhageHostLearn

    # All datasets:
    python scripts/analysis/compare_eval_methodologies.py <exp_dir>

    # Write also as .xlsx (one sheet per dataset) if openpyxl is installed:
    python scripts/analysis/compare_eval_methodologies.py <exp_dir> --xlsx
"""

import argparse
import csv
import json
import os
import sys
from collections import defaultdict
from pathlib import Path

import numpy as np
import yaml


# Make `cipher` importable from a checkout without `pip install -e .`
def _ensure_cipher_on_path():
    here = Path(__file__).resolve().parent
    for parent in [here, *here.parents]:
        candidate = parent / 'src' / 'cipher'
        if candidate.is_dir():
            src = str(parent / 'src')
            if src not in sys.path:
                sys.path.insert(0, src)
            return

_ensure_cipher_on_path()

from cipher.data.interactions import (
    load_interaction_pairs, load_phage_protein_mapping,
)
from cipher.evaluation.ranking import rank_hosts, _ranks_with_ties
from cipher.evaluation.runner import load_predictor, load_validation_data


def _resolve_val_paths(experiment_dir, cli):
    cfg_path = os.path.join(experiment_dir, 'config.yaml')
    val_cfg = {}
    if os.path.exists(cfg_path):
        with open(cfg_path) as f:
            val_cfg = yaml.safe_load(f).get('validation', {}) or {}
    return {
        'fasta': cli.val_fasta or val_cfg.get('val_fasta'),
        'emb': cli.val_embedding_file or val_cfg.get('val_embedding_file'),
        'emb2': cli.val_embedding_file_2 or val_cfg.get('val_embedding_file_2'),
        'ds_dir': cli.val_datasets_dir or val_cfg.get('val_datasets_dir'),
    }


def precompute_protein_probs(predictor, md5s, emb_dict):
    """Return (md5_to_idx, k_probs[N,Kclasses], o_probs[N,Oclasses])."""
    k_classes = list(predictor.k_classes)
    o_classes = list(predictor.o_classes)
    k_idx = {c: i for i, c in enumerate(k_classes)}
    o_idx = {c: i for i, c in enumerate(o_classes)}

    rows_md5 = []
    k_rows = []
    o_rows = []
    for m in md5s:
        if m not in emb_dict:
            continue
        out = predictor.predict_protein(emb_dict[m])
        kp = np.zeros(len(k_classes), dtype=np.float32)
        op = np.zeros(len(o_classes), dtype=np.float32)
        for c, p in out['k_probs'].items():
            if c in k_idx:
                kp[k_idx[c]] = float(p)
        for c, p in out['o_probs'].items():
            if c in o_idx:
                op[o_idx[c]] = float(p)
        rows_md5.append(m)
        k_rows.append(kp)
        o_rows.append(op)

    if not rows_md5:
        return {}, np.zeros((0, len(k_classes)), dtype=np.float32), \
               np.zeros((0, len(o_classes)), dtype=np.float32), k_classes, o_classes
    return ({m: i for i, m in enumerate(rows_md5)},
            np.stack(k_rows), np.stack(o_rows), k_classes, o_classes)


def old_eval_for_pair(proteins, true_k, true_o, pid_md5, md5_to_idx,
                      k_probs, o_probs, k_classes, o_classes):
    """Replicate OLD eval's `--merge-strategy raw` behavior for one pair.

    Returns dict or None if no usable RBP.
    """
    has_k = true_k not in ('N/A', 'null', 'Unknown', '')
    has_o = true_o not in ('N/A', 'null', 'Unknown', '')

    best_merged_rank = None
    best_top1 = None       # (class, prob, head) for best RBP's overall top-1
    best_rbp = None
    best_k_rank = None
    best_o_rank = None

    for pid in proteins:
        md5 = pid_md5.get(pid)
        if md5 is None or md5 not in md5_to_idx:
            continue
        idx = md5_to_idx[md5]

        # Build merged class list: K classes + O classes, sorted by prob desc.
        # Tie-breaking: stable, K-first (matches numpy argsort + Python sort
        # used in the old script: sort uses key=-prob, stable on original order).
        combined = []
        for j, c in enumerate(k_classes):
            combined.append((c, float(k_probs[idx, j]), 'K'))
        for j, c in enumerate(o_classes):
            combined.append((c, float(o_probs[idx, j]), 'O'))
        combined.sort(key=lambda x: -x[1])

        merged_rank = None
        for r, (c, _p, _h) in enumerate(combined, 1):
            if (has_k and c == true_k) or (has_o and c == true_o):
                merged_rank = r
                break

        if merged_rank is not None and (best_merged_rank is None
                                        or merged_rank < best_merged_rank):
            best_merged_rank = merged_rank
            best_top1 = combined[0]
            best_rbp = pid

        if has_k:
            ksort = np.argsort(-k_probs[idx])
            for r, ci in enumerate(ksort, 1):
                if k_classes[ci] == true_k:
                    if best_k_rank is None or r < best_k_rank:
                        best_k_rank = r
                    break
        if has_o:
            osort = np.argsort(-o_probs[idx])
            for r, ci in enumerate(osort, 1):
                if o_classes[ci] == true_o:
                    if best_o_rank is None or r < best_o_rank:
                        best_o_rank = r
                    break

    if best_merged_rank is None or best_top1 is None:
        return None

    top1_class, top1_prob, top1_head = best_top1
    return {
        'merged_rank': best_merged_rank,
        'top1_class': top1_class,
        'top1_prob': top1_prob,
        'top1_head': top1_head,
        'top1_correct': (has_k and top1_class == true_k) or
                        (has_o and top1_class == true_o),
        'best_rbp': best_rbp,
        'k_rank': best_k_rank,
        'o_rank': best_o_rank,
    }


COLUMNS = [
    'dataset', 'phage_id', 'host_id',
    'true_K', 'true_O', 'n_rbps_with_emb',
    # OLD class-ranking (raw merge)
    'old_merged_rank',
    'old_top1_class', 'old_top1_prob', 'old_top1_head',
    'old_top1_correct',
    'old_best_rbp',
    'old_k_rank', 'old_o_rank',
    # NEW host-ranking
    'new_host_rank',
    'new_top1_host', 'new_top1_host_K', 'new_top1_host_O',
    'new_top1_score',
    'new_top1_correct',
]


def run_for_dataset(predictor, dataset_dir, dataset_name,
                    emb_dict, pid_md5):
    """Run both eval methodologies and return (rows, summary).

    `rows` is a list of dicts keyed by COLUMNS, with native Python types
    (int / float / bool / str / None) so that downstream writers can
    emit either CSV or properly-typed xlsx cells.
    """
    pairs = load_interaction_pairs(dataset_dir)
    pm_path = os.path.join(dataset_dir, 'metadata', 'phage_protein_mapping.csv')
    phage_protein_map = load_phage_protein_mapping(pm_path)

    interactions = defaultdict(dict)
    serotypes = {}
    for p in pairs:
        interactions[p['phage_id']][p['host_id']] = p['label']
        serotypes[p['host_id']] = {'K': p['host_K'], 'O': p['host_O']}

    needed_md5s = set()
    for pid_set in phage_protein_map.values():
        for pid in pid_set:
            m = pid_md5.get(pid)
            if m is not None and m in emb_dict:
                needed_md5s.add(m)
    md5_to_idx, k_probs_mat, o_probs_mat, k_classes, o_classes = \
        precompute_protein_probs(predictor, sorted(needed_md5s), emb_dict)

    rows = []
    n_old_top1 = n_old_top5 = n_new_top1 = n_new_top5 = n_agree_top1 = 0

    for phage in sorted(interactions.keys()):
        pos_hosts = [h for h, lbl in interactions[phage].items() if lbl == 1]
        if not pos_hosts:
            continue

        proteins = phage_protein_map.get(phage, set())
        n_proteins_with_emb = sum(
            1 for pid in proteins if pid_md5.get(pid) in md5_to_idx
        )
        if n_proteins_with_emb == 0:
            continue

        candidates = list(interactions[phage].keys())
        ranked = rank_hosts(predictor, proteins, candidates, serotypes,
                            emb_dict, pid_md5)
        host_to_rank = _ranks_with_ties(ranked, tie_method='competition')
        top1_host_id, top1_score = ranked[0] if ranked else (None, None)
        top1_host_K = serotypes.get(top1_host_id, {}).get('K') if top1_host_id else None
        top1_host_O = serotypes.get(top1_host_id, {}).get('O') if top1_host_id else None

        for pos_h in pos_hosts:
            if pos_h not in serotypes:
                continue
            true_k = serotypes[pos_h]['K']
            true_o = serotypes[pos_h]['O']

            old = old_eval_for_pair(
                proteins, true_k, true_o, pid_md5,
                md5_to_idx, k_probs_mat, o_probs_mat,
                k_classes, o_classes,
            )
            new_rank = host_to_rank.get(pos_h)
            new_correct = (top1_host_id == pos_h)

            if old is None and new_rank is None:
                continue

            rows.append({
                'dataset': dataset_name,
                'phage_id': phage,
                'host_id': pos_h,
                'true_K': true_k,
                'true_O': true_o,
                'n_rbps_with_emb': int(n_proteins_with_emb),
                'old_merged_rank': int(old['merged_rank']) if old else None,
                'old_top1_class': old['top1_class'] if old else None,
                'old_top1_prob': float(old['top1_prob']) if old else None,
                'old_top1_head': old['top1_head'] if old else None,
                'old_top1_correct': bool(old['top1_correct']) if old else None,
                'old_best_rbp': old['best_rbp'] if old else None,
                'old_k_rank': int(old['k_rank']) if old and old['k_rank'] is not None else None,
                'old_o_rank': int(old['o_rank']) if old and old['o_rank'] is not None else None,
                'new_host_rank': int(new_rank) if new_rank is not None else None,
                'new_top1_host': top1_host_id,
                'new_top1_host_K': top1_host_K,
                'new_top1_host_O': top1_host_O,
                'new_top1_score': float(top1_score) if top1_score is not None else None,
                'new_top1_correct': bool(new_correct),
            })

            if old is not None:
                if old['merged_rank'] <= 1: n_old_top1 += 1
                if old['merged_rank'] <= 5: n_old_top5 += 1
            if new_rank is not None:
                if new_rank <= 1: n_new_top1 += 1
                if new_rank <= 5: n_new_top5 += 1
            if (old is not None and old['merged_rank'] <= 1
                    and new_rank is not None and new_rank <= 1):
                n_agree_top1 += 1

    n = max(len(rows), 1)
    summary = {
        'dataset': dataset_name,
        'n_pairs': len(rows),
        'old_HR@1': n_old_top1 / n,
        'old_HR@5': n_old_top5 / n,
        'new_HR@1': n_new_top1 / n,
        'new_HR@5': n_new_top5 / n,
        'both_top1': n_agree_top1 / n,
    }
    return rows, summary


def write_dataset_csv(rows, out_csv):
    os.makedirs(os.path.dirname(out_csv) or '.', exist_ok=True)
    with open(out_csv, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(COLUMNS)
        for r in rows:
            row_out = []
            for col in COLUMNS:
                v = r.get(col)
                if v is None:
                    row_out.append('')
                elif isinstance(v, float):
                    row_out.append(f'{v:.4f}')
                elif isinstance(v, bool):
                    row_out.append(str(v))
                else:
                    row_out.append(v)
            w.writerow(row_out)


def write_xlsx(per_dataset, summaries, xlsx_path):
    """Write a single workbook with typed cells, freeze pane, autofilter."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('summary')
    ws.append(['dataset', 'n_pairs',
               'OLD HR@1', 'OLD HR@5',
               'NEW HR@1', 'NEW HR@5',
               'both top1'])
    for s in summaries:
        ws.append([s['dataset'], int(s['n_pairs']),
                   round(s['old_HR@1'], 4), round(s['old_HR@5'], 4),
                   round(s['new_HR@1'], 4), round(s['new_HR@5'], 4),
                   round(s['both_top1'], 4)])
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    for ds_name, rows in per_dataset:
        ws = wb.create_sheet(ds_name[:31])
        ws.append(COLUMNS)
        for r in rows:
            ws.append([r.get(c) for c in COLUMNS])
        ws.freeze_panes = 'A2'
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
        # Column widths: clip max length per column to keep file small
        for ci, col in enumerate(COLUMNS, start=1):
            max_len = max(
                [len(col)] + [len(str(r.get(col)) if r.get(col) is not None else '')
                              for r in rows]
            )
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 32)

    wb.save(xlsx_path)


def main():
    p = argparse.ArgumentParser()
    p.add_argument('experiment_dir',
                   help='Path to a trained experiment directory (model_k/, model_o/)')
    p.add_argument('--datasets', nargs='+', default=None,
                   help='Subset of validation datasets (default: all available)')
    p.add_argument('--out-dir', default=None,
                   help='Default: <experiment_dir>/results/analysis/')
    p.add_argument('--val-fasta', default=None)
    p.add_argument('--val-embedding-file', default=None)
    p.add_argument('--val-embedding-file-2', default=None)
    p.add_argument('--val-datasets-dir', default=None)
    p.add_argument('--no-xlsx', action='store_true',
                   help='Skip the .xlsx workbook (CSVs always written)')
    args = p.parse_args()

    if not os.path.isdir(args.experiment_dir):
        sys.exit(f'ERROR: {args.experiment_dir} is not a directory')

    paths = _resolve_val_paths(args.experiment_dir, args)
    for key in ('fasta', 'emb', 'ds_dir'):
        if not paths.get(key):
            sys.exit(f'ERROR: missing validation path {key!r}; '
                     f'pass via CLI or put in config.yaml:validation')

    out_dir = args.out_dir or os.path.join(
        args.experiment_dir, 'results', 'analysis')
    os.makedirs(out_dir, exist_ok=True)

    print(f'Experiment: {args.experiment_dir}')
    print(f'Out dir:    {out_dir}')

    print('Loading predictor and validation data ...')
    predictor = load_predictor(args.experiment_dir)
    val_data = load_validation_data(
        paths['fasta'], paths['emb'], paths['ds_dir'],
        val_embedding_file_2=paths['emb2'])

    datasets = args.datasets or val_data['available_datasets']
    per_dataset = []
    summaries = []
    for ds in datasets:
        ds_dir = os.path.join(paths['ds_dir'], ds)
        if not os.path.isdir(ds_dir):
            print(f'  Skipping {ds} (not found at {ds_dir})')
            continue
        print(f'  {ds}: running both eval methodologies')
        rows, summary = run_for_dataset(
            predictor, ds_dir, ds,
            val_data['emb_dict'], val_data['pid_md5'])
        out_csv = os.path.join(out_dir, f'{ds}_old_vs_new_eval.csv')
        write_dataset_csv(rows, out_csv)
        per_dataset.append((ds, rows))
        summaries.append(summary)

    print()
    print('=' * 100)
    print(f'{"dataset":<18} {"n":>6} {"OLD HR@1":>10} {"OLD HR@5":>10} '
          f'{"NEW HR@1":>10} {"NEW HR@5":>10} {"both top1":>10}')
    print('-' * 100)
    for s in summaries:
        print(f'{s["dataset"]:<18} {s["n_pairs"]:>6} '
              f'{s["old_HR@1"]:>10.4f} {s["old_HR@5"]:>10.4f} '
              f'{s["new_HR@1"]:>10.4f} {s["new_HR@5"]:>10.4f} '
              f'{s["both_top1"]:>10.4f}')

    summary_path = os.path.join(out_dir, 'old_vs_new_eval_summary.json')
    with open(summary_path, 'w') as f:
        json.dump(summaries, f, indent=2)
    print(f'\nSummary: {summary_path}')

    if not args.no_xlsx:
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            print('WARNING: openpyxl not installed, skipping xlsx output')
        else:
            xlsx_path = os.path.join(out_dir, 'old_vs_new_eval.xlsx')
            write_xlsx(per_dataset, summaries, xlsx_path)
            print(f'Workbook: {xlsx_path}')


if __name__ == '__main__':
    main()
